
from flask import send_file
import shutil
import openpyxl
from openpyxl import load_workbook
import time


class DownloadController:

    def __init__(self,parameterService,tickerRateService):
        self.parameterService = parameterService
        self.tickerRateService = tickerRateService
        pass

    def dispatch(self, request):
        tickers, from_date, till_date = self.parameterService.init_params(1500)
        tickers = "goog"

        ticker_data = self.tickerRateService.get_rate(tickers, from_date, till_date)

        dest_file = 'static/excel/iteration8copy.xlsx'
        shutil.copy('static/excel/iteration8.xlsx', dest_file)


        wb = load_workbook(filename=dest_file)
        ws = wb["Analysis"]
        ws["F7"] = tickers
        ws["F6"] = ticker_data.iloc[0]['Close']

        #ws["F6"] = from_date
        #ws["b6"] = till_date
        #ws["d4"] = ticker_data[0]["Close"]
        wb.save(dest_file)
        print(time.time())
        result = send_file(dest_file,
                         mimetype='text/csv',
                         attachment_filename='dummy.xlsm',
                         as_attachment=True)
        print(time.time())
        return result



